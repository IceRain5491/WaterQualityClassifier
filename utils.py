import csv, os
from typing import Tuple, Dict, List
import pandas as pd
import numpy as np
import re
import unicodedata

from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import PatternFill
from PIL import Image as PILImage
from io import BytesIO


# —— 将类似“-- / 未检出 / NA / 0.12mg/L / <0.05”规范为数值 ——
MISSING_TOKENS = {"", "--", "—", "-", "–", "－", "nan", "none", "null", "na", "n/a", "nd", "n.d.", "未检出", "空", "无"}

def _coerce_series_to_numeric(s):
    
    if pd.api.types.is_numeric_dtype(s):
        return s
    s2 = s.astype(str).str.strip().str.lower()
    s2 = s2.replace({tok: None for tok in MISSING_TOKENS})
    s2 = s2.str.replace(r"[<>≤≥≈~～\s]", "", regex=True)
    s2 = s2.str.replace(r"[^0-9.\-eE]", "", regex=True)
    return pd.to_numeric(s2, errors="coerce")

def normalize_numeric_columns(df, exclude_cols):
    df2 = df.copy()
    for col in df2.columns:
        if col in exclude_cols:
            continue
        try:
            df2[col] = _coerce_series_to_numeric(df2[col])
        except Exception:
            pass
    return df2

def resample_df(df: pd.DataFrame, date_col: str, freq_code: str | None) -> pd.DataFrame:
    data = df.copy()
    data[date_col] = pd.to_datetime(data[date_col], errors='coerce')
    data = data.dropna(subset=[date_col]).sort_values(by=date_col).set_index(date_col)
    if freq_code:
        num = data.select_dtypes(include='number').columns
        data = data[num].resample(freq_code).mean().dropna(how='all').reset_index()
    else:
        data = data.reset_index()
    return data

def infer_freq_from_index(dti: pd.DatetimeIndex) -> str | None:
    if len(dti) < 3:
        return None
    try:
        f = pd.infer_freq(dti)
    except Exception:
        f = None
    if f:
        u = f.upper()
        for k in ('H','D','W','M','Q','A','Y'):
            if u.startswith(k): return 'A' if k=='Y' else k
    deltas = np.diff(dti.asi8)
    if len(deltas)==0: return None
    median_ns = np.median(deltas)
    day_ns = 24*3600*1e9
    if median_ns >= 25*day_ns: return 'M'
    if median_ns >= day_ns: return 'D'
    if median_ns >= day_ns/24: return 'H'
    return None

def choose_numeric_datefmt(t: pd.Series) -> str:
    """根据时间跨度选择数字日期格式，避免英文月份."""
    t = pd.to_datetime(t)
    if t.empty:
        return "%Y-%m-%d"
    span = (t.max() - t.min()).to_pytimedelta().days if hasattr((t.max() - t.min()), 'to_pytimedelta') else (t.max() - t.min()).days
    if span >= 365:
        return "%Y-%m"
    if span >= 60:
        return "%Y-%m"
    if span >= 7:
        return "%m-%d"
    if span >= 1:
        return "%m-%d"
    return "%H:%M"


def normalize_category(raw: str) -> str:
    if raw is None:
        return ""
    s = str(raw).strip()
    if not s:
        return ""

    s = re.sub(r"\s+", "", s)
    roman_map = {"Ⅰ": "I", "Ⅱ": "II", "Ⅲ": "III", "Ⅳ": "IV", "Ⅴ": "V"}
    for k, v in roman_map.items():
        s = s.replace(k, v)
    if s.endswith("类"):
        s = s[:-1]
    if s.startswith("劣"):
        return "劣V类"
    digit2roman = {"1": "I", "2": "II", "3": "III", "4": "IV", "5": "V"}
    s = digit2roman.get(s, s)
    return f"{s}类" if s in {"I","II","III","IV","V"} else ""


def hex_to_openpyxl(hexcolor: str) -> str:
    """把 '#rrggbb' 或 'rrggbb' 转为 openpyxl 可以用的 'RRGGBB'（大写，不带#）。"""
    if not hexcolor:
        return None
    s = str(hexcolor).lstrip('#').strip()
    if len(s) == 3:  # shorthand e.g. 'f00'
        s = ''.join([c*2 for c in s])
    s = s.upper()
    if len(s) != 6:
        return None
    return s

def calculate_display_width(text: str) -> int:
    """中文/全角字符宽度为2，其它为1。用于 Excel 列宽估算。"""
    if text is None:
        return 0
    width = 0
    for ch in str(text):
        if unicodedata.east_asian_width(ch) in ('F', 'W'):
            width += 2
        else:
            width += 1
    return width

def sanitize_label(s: str) -> str:
    """NFKC 规范化，去零宽/控制字符，合并空白并 trim。保证拼 key 不会含不可见字符。"""
    if s is None:
        return ""
    t = unicodedata.normalize("NFKC", str(s))
    # 去掉常见零宽字符
    t = re.sub(r"[\u200B-\u200F\uFEFF\u2060]", "", t)
    # 把连续空白合并成一个空格，再去首尾空格
    t = re.sub(r"\s+", " ", t).strip()
    return t

def load_station_mapping(csv_path: str, encoding_list=("utf-8", "gbk", "cp936")) -> Tuple[Dict[str, Tuple[str,str,str,str]], Dict[str, List[str]]]:
    """
    更鲁棒的 CSV 读取：尝试若干编码并对常见列名做容错匹配。
    返回 (station_coords, station_groups)。
    """
    station_coords = {}
    station_groups = {}
    if not os.path.exists(csv_path):
        alt = os.path.join(os.path.dirname(__file__), os.path.basename(csv_path))
        if os.path.exists(alt):
            csv_path = alt
        else:
            print(f"[wq_utils] load_station_mapping: not found: {csv_path}")
            return station_coords, station_groups

    col_name_candidates = {
        "name": ["#名称", "名称", "name", "站名", "站点", "点位", "site", "station"],
        "lon":  ["经度", "lon", "longitude"],
        "lat":  ["纬度", "lat", "latitude"],
        "wtype":["水体类型", "type", "water_type"],
        "group":["点位分组", "分组", "group"]
    }

    for enc in encoding_list:
        try:
            with open(csv_path, encoding=enc, errors="replace") as f:
                reader = csv.DictReader(f)
                fieldnames = reader.fieldnames or []
                header_map = {sanitize_label(h): h for h in fieldnames}
                for row in reader:
                    def pick(cands):
                        for k in cands:
                            k2 = sanitize_label(k)
                            if k2 in header_map:
                                return (row.get(header_map[k2]) or "").strip()
                        for k in cands:
                            if k in row:
                                return (row.get(k) or "").strip()
                        return ""
                    name = pick(col_name_candidates["name"])
                    lon  = pick(col_name_candidates["lon"])
                    lat  = pick(col_name_candidates["lat"])
                    wtype= pick(col_name_candidates["wtype"])
                    group= pick(col_name_candidates["group"]) or ""
                    if name:
                        name = sanitize_label(name)
                        station_coords[name] = (lon, lat, wtype, group)
                        station_groups.setdefault(group, []).append(name)
            return station_coords, station_groups
        except Exception as e:
            print(f"[wq_utils] try encoding {enc} failed: {e}")
            continue

    print(f"[wq_utils] failed to read CSV with tried encodings: {encoding_list}")
    return {}, {}
    
def build_row_data(station_name: str, sampling_date: str,
                   metric_definitions: list, include_categories: bool,
                   include_lonlat: bool, lon: str="", lat: str="") -> dict:
    """
    metric_definitions: [(name, val, cat), ...]
    返回规范化的 row_data 字典（所有键与值都会 sanitize）
    """
    row_data = {}
    row_data["点位名称"] = sanitize_label(station_name)
    row_data["监测日期"] = sanitize_label(sampling_date)

    for name, val, cat in metric_definitions:
        if val is None or str(val).strip() == "":
            continue
        key_name = sanitize_label(name)
        row_data[key_name] = str(val).strip()
        if include_categories:
            # **注意**：这里严格拼接为 f"{name}类别"（没有任何空格）
            cat_key = sanitize_label(f"{name}类别")
            row_data[cat_key] = sanitize_label(cat)

    if include_lonlat and lon and lat:
        row_data["经度"] = sanitize_label(lon)
        row_data["纬度"] = sanitize_label(lat)

    groups = {"IV类": [], "V类": [], "劣V类": []}
    for name, val, cat in metric_definitions:
        if val and cat in groups:
            groups[cat].append(f"{name}（{val}）")
    row_data["四类因子"] = "，".join(groups["IV类"])
    row_data["五类因子"] = "，".join(groups["V类"])
    row_data["劣五类因子"] = "，".join(groups["劣V类"])

    return row_data

def ensure_headers(ws, desired_headers: list):
    """保证 ws 第一行包含 desired_headers（按顺序插入缺失列），返回最终 headers 列表"""
    current = [cell.value for cell in ws[1]]
    # 若 ws 第一行为空（新表格），直接写入
    if not any(current):
        for i, h in enumerate(desired_headers, start=1):
            ws.cell(row=1, column=i, value=h)
        return desired_headers[:]
    # 否则按索引插入缺失项（保持顺序）
    cur = current[:]  # copy mutable
    for i, h in enumerate(desired_headers):
        if h not in cur:
            # 插入到 i 位置
            ws.insert_cols(i+1)
            ws.cell(row=1, column=i+1, value=h)
            cur.insert(i, h)
    return cur

def append_row_with_headers(ws, desired_headers: list, row_data: dict):
    """确保表头并追加 row_data（按 desired_headers 顺序）。返回 (headers_after, new_row_rowidx)."""
    headers_after = ensure_headers(ws, desired_headers)
    row = [row_data.get(h, "") for h in headers_after]
    ws.append(row)
    return headers_after, ws.max_row

def auto_adjust_column_width(ws, headers: list, calc_width_func):
    """headers: 列头列表，calc_width_func(text)->int"""
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        maxw = calc_width_func(header)
        for cell in ws[col_letter]:
            maxw = max(maxw, calc_width_func(str(cell.value or "")))
        ws.column_dimensions[col_letter].width = maxw + 2

def insert_image_to_cell(ws, img_path: str, row:int, col:int, max_edge=300):
    if not img_path or not os.path.exists(img_path):
        return False
    try:
        pil = PILImage.open(img_path)
        w0, h0 = pil.size
        scale = min(max_edge / max(w0, h0), 1)
        nw, nh = int(w0 * scale), int(h0 * scale)
        pil = pil.resize((nw, nh), PILImage.LANCZOS)
        bio = BytesIO()
        pil.save(bio, format="PNG")
        bio.seek(0)
        img = XLImage(bio)
        img.width, img.height = nw, nh
        cell_ref = f"{get_column_letter(col)}{row}"
        ws.add_image(img, cell_ref)
        ws.row_dimensions[row].height = nh * 0.75
        return True
    except Exception:
        return False
    
def canonical_metric_id(name: str) -> str | None:
    n = (name or "").lower()
    if "ph" in n: return "pH"
    if "溶解氧" in n or "do" in n: return "溶解氧"
    if "高锰酸" in n or "cod" in n or "codmn" in n: return "CODMn"
    if "氨氮" in n or "nh" in n: return "氨氮"
    if "总磷" in n or "tp" in n: return "总磷"
    return None

def _normalize_cat(cat: str, classifier_cls) -> str:
    """
    把各种写法（'合格'、'I类'、'Ⅰ'、'1' 等）规范到标准类别名（如 'I类','II类',...,'劣V类'）。
    优先使用 classifier_cls.NORMALIZE_CATEGORY（如果存在）。
    fallback: 简单清洗 + 数字→罗马转换 + '劣' 前缀处理。
    """
    if not cat:
        return ""
    s = str(cat).strip()
    # 优先用类里定义的映射（若有）
    mapping = getattr(classifier_cls, "NORMALIZE_CATEGORY", None)
    if mapping and s in mapping:
        return mapping[s]

    # 尝试更宽松的匹配（去空白、规范化）
    try:
        s2 = sanitize_label(s)
    except Exception:
        s2 = s

    if mapping and s2 in mapping:
        return mapping[s2]

    # 去掉 '类' 后缀
    if s2.endswith("类"):
        core = s2[:-1]
    else:
        core = s2

    # 罗马数字全角→ASCII already handled by sanitize_label in many cases; 再处理几种常见
    roman_map = {"Ⅰ":"I","Ⅱ":"II","Ⅲ":"III","Ⅳ":"IV","Ⅴ":"V"}
    for k,v in roman_map.items():
        core = core.replace(k, v)

    # 数字转罗马
    digit2roman = {"1":"I","2":"II","3":"III","4":"IV","5":"V"}
    if core in digit2roman:
        core = digit2roman[core]

    # 劣V类开头
    if core.startswith("劣"):
        return "劣V类"

    # 如果已经是 I/II/... 则补上 '类'
    if core in {"I","II","III","IV","V"}:
        return f"{core}类"

    # 处理“合格”等常见词（尝试再查 mapping）
    if core.lower() in {"合格", "qualified", "pass"} and mapping:
        return mapping.get("合格", "I类")

    # fallback: 返回原始（让调用方处理空或未知）
    return s


def compute_overall_from_categories(categories, classifier_cls):
    """
    categories: 可含各种写法（'合格','I类','Ⅱ类' 等）
    classifier_cls: WaterQualityClassifier 类（提供 CATEGORY_ORDER, ORDER_TO_CATEGORY, 可能的 NORMALIZE_CATEGORY）
    返回规范化后的总体类别（字符串），无法判断则返回 "无有效数据" 或 ""（与早期行为一致）。
    """
    if not categories:
        return "无有效数据"
    # 规范化后按 order 取最差
    orders = []
    for c in categories:
        norm = _normalize_cat(c, classifier_cls)
        order = classifier_cls.CATEGORY_ORDER.get(norm, None)
        if order is None:
            # 有可能 mapping 返回空或未命中，尝试直接用未规范化值再查一次
            order = classifier_cls.CATEGORY_ORDER.get(c, 0)
        orders.append(order or 0)
    worst = max(orders) if orders else 0
    return classifier_cls.ORDER_TO_CATEGORY.get(worst, "")


def color_category_cells(ws, row: int, headers: list, classifier_cls):
    """
    为行 `row` 中所有以 '类别' 结尾的列上色，使用 classifier_cls.CATEGORY_COLORS。
    headers: 表头列表（与工作表第一行对应）
    classifier_cls: WaterQualityClassifier class
    """
    if not headers:
        headers = [cell.value for cell in ws[1]]

    for col_idx, header in enumerate(headers, start=1):
        if not header:
            continue
        # 只处理以 '类别' 结尾的列
        if str(header).endswith("类别"):
            cell = ws.cell(row=row, column=col_idx)
            val = cell.value
            if not val:
                continue

            # 先尝试把单元格值正规化为标准类别文本（例如把 "合格" -> "I类"）
            try:
                norm_val = normalize_category(val)
            except Exception:
                norm_val = ""

            color_hex = None
            if norm_val:
                color_hex = classifier_cls.CATEGORY_COLORS.get(norm_val)

            # 仍然找不到时，尝试直接按原值寻找（兼容直接写了 'I类' 的情况）
            if color_hex is None:
                color_hex = classifier_cls.CATEGORY_COLORS.get(str(val))

            # 最终备用：按 header 基本指标名查色（例如 'pH类别' -> 'pH'）
            if color_hex is None:
                base = header[:-2] if header.endswith("类别") else header
                color_hex = classifier_cls.CATEGORY_COLORS.get(base)

            if not color_hex:
                continue

            rgb = hex_to_openpyxl(color_hex)
            if not rgb:
                continue

            # 用 ARGB 的形式填色（加上 FF 以保证不透明）
            try:
                fill = PatternFill(fill_type='solid', start_color="FF"+rgb, end_color="FF"+rgb)
                cell.fill = fill
            except Exception:
                try:
                    fill = PatternFill(fill_type='solid', start_color=rgb, end_color=rgb)
                    cell.fill = fill
                except Exception:
                    pass

